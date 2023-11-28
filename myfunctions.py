from pathlib import Path

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service

# Path to project root:
ROOT_FOLDER = Path(__file__).parent
# Path to chromedriver:
CHROOMEDRIVER_EXEC = ROOT_FOLDER / 'drivers' / 'chromedriver.exe'
# Chrome binary path- version compatible with chromedriver:
CHROME_BINARY_PATH = ROOT_FOLDER / 'chrome-win32' / 'chrome.exe'

# Function to read dataset from excel


def load_excel_data(file_path):
    workbook = load_workbook(file_path)
    worksheet = workbook['webservices']
    return [row[0].value for row in worksheet.iter_rows(min_row=2, min_col=2,
            max_col=2) if row[0].value is not None]

# Function to open Chrome


def make_chrome_browser(*options: str) -> webdriver.Chrome:
    chrome_options = webdriver.ChromeOptions()

    # chrome_options.add_argument('--headless')
    if options is not None:
        for option in options:
            chrome_options.add_argument(option)  # # type: ignore (se tiver
            # problemas com tipagem)

    # Convertendo o Path para string
    chrome_options.binary_location = CHROME_BINARY_PATH.as_posix()
    chrome_service = Service(
        executable_path=str(CHROOMEDRIVER_EXEC),
    )

    chrome_browser = webdriver.Chrome(
        service=chrome_service,
        options=chrome_options,
    )

    return chrome_browser

# Function to open tabs in Chrome


def open_webservices(chrome_browser, *list_links):
    for i, n in enumerate(list_links):
        try:
            # Open a new browser tab with URL 'n':
            chrome_browser.execute_script(f"window.open('{n}', 'tab+{i}');")
            # Switch to the last open tab.
            chrome_browser.switch_to.window(chrome_browser.window_handles[-1])
            print(i, n)
        except Exception as e:
            print(f"Error opening tab {i}: {e}")
    return
